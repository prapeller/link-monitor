import fastapi as fa
import openpyxl
from pydantic.error_wrappers import ValidationError

from core.exceptions import NotExcelException, WhileUploadingArchiveException
from database import SessionLocal
from database.crud import get_or_create, get, create_user_from_keycloak
from database.models.link_url_domain import LinkUrlDomainModel
from database.models.page_url_domain import PageUrlDomainModel
from database.models.user import UserModel
from database.schemas.link import LinkCreateWithDomainsSerializer
from database.schemas.link_url_domain import LinkUrlDomainCreateSerializer
from database.schemas.page_url_domain import PageUrlDomainCreateSerializer
from services.keycloak import KCAdmin
from core.shared import get_domain_name_from_url


def iterfile(filepath):
    with open(filepath, mode="rb") as file:
        yield from file


def get_duplicated_link_row_numbers(ws_links: list[tuple], duplicated_link: tuple) -> str:
    """check if link_tuple(link_url, anchor, page_url) has duplicates in list of such tuples (ws_link_list)"""
    row_numbers = []

    for row_num, link in enumerate(ws_links, start=1):
        if link == duplicated_link:
            row_numbers.append(str(row_num))

    return ', '.join(row_numbers)


def check_file_on_duplicates(filepath) -> list[str]:
    """
    check if file has duplicated links (violating unique constraint (link_url, anchor, page_url))
    and returns list of 'error strings' like
    [
        "file has duplicated links at rows 1, 10",
        "file has duplicated links at rows 2, 12",
    ]
    """
    try:
        wb = openpyxl.load_workbook(filename=filepath)
    except KeyError:
        raise NotExcelException
    ws = wb.active

    ws_links: list[tuple] = []
    ws_duplicated_links: list[tuple] = []
    validation_errors = []

    for row in ws.iter_rows():
        created_at, link_url, anchor, page_url = [cell.value for cell in row[:4]]
        if all((created_at, link_url, anchor, page_url)):
            if not link_url.endswith('/'):
                link_url = link_url + '/'
            ws_links.append((link_url, anchor, page_url))
        else:
            continue

    for link in ws_links:
        if ws_links.count(link) > 1:
            ws_duplicated_links.append(link)
            print(f'found duplicate link: {link}')

    if ws_duplicated_links:
        ws_duplicated_links = list(set(ws_duplicated_links))
        print(f'duplicated links: {ws_duplicated_links}')

        for duplicated_link in ws_duplicated_links:
            row_numbers = get_duplicated_link_row_numbers(ws_links, duplicated_link)
            error_string = f'error at row numbers {row_numbers}: link with ' \
                           f'link_url={duplicated_link[0]}, ' \
                           f'anchor={duplicated_link[1]}, ' \
                           f'page_url={duplicated_link[2]} ' \
                           f'is duplicated in file'
            print(error_string)
            validation_errors.append(error_string)

    return validation_errors


def get_link_create_ser_list_from_file(filepath, current_user_id: int, mode=None) \
        -> list[LinkCreateWithDomainsSerializer]:
    db = SessionLocal()
    current_user = get(db, UserModel, id=current_user_id)

    try:
        wb = openpyxl.load_workbook(filename=filepath)
    except KeyError:
        raise NotExcelException
    ws = wb.active

    link_create_ser_list = []
    validation_errors = []

    if mode is None:

        for num, row in enumerate(ws.iter_rows(), start=1):
            if num == 1:
                continue

            row_list = [cell.value for cell in row[:8]]
            if all((cell is None for cell in row_list)):
                continue

            page_url, link_url, anchor, da, dr, price, screenshot_url, contact = row_list
            link_url, anchor, page_url = str(link_url).strip(), str(anchor).strip(), str(page_url).strip()

            if not all((page_url, link_url, anchor, da, dr, price, contact)):
                validation_errors.append(
                    f'error at row number {num}: there are empty cells, only screenshot_url can be empty')
                continue

            if not link_url.endswith('/'):
                link_url = link_url + '/'

            link_url_domain_name = get_domain_name_from_url(link_url)
            link_url_domain_ser = LinkUrlDomainCreateSerializer(name=link_url_domain_name)
            link_url_domain = get_or_create(db, LinkUrlDomainModel, link_url_domain_ser)

            page_url_domain_name = get_domain_name_from_url(page_url)
            page_url_domain_ser = PageUrlDomainCreateSerializer(name=page_url_domain_name)
            page_url_domain = get_or_create(db, PageUrlDomainModel, page_url_domain_ser)

            link_data = {
                'user_id': current_user_id,
                'page_url': page_url,
                'link_url': link_url,
                'anchor': anchor,
                'da': da,
                'dr': dr,
                'price': price,
                'screenshot_url': screenshot_url,
                'contact': contact,
                'link_url_domain_id': link_url_domain.id,
                'page_url_domain_id': page_url_domain.id,
            }
            try:
                link_create_ser = LinkCreateWithDomainsSerializer(**link_data)
                link_create_ser_list.append(link_create_ser)
            except ValidationError as e:
                error = f'error at row number {num}: {e}'
                validation_errors.append(error)
                print(error)

    elif mode == 'from_archive':
        kc_admin = KCAdmin()

        for num, row in enumerate(ws.iter_rows(), start=1):
            if num == 1:
                continue

            row_list = [cell.value for cell in row[:11]]
            if all((cell is None for cell in row_list)):
                continue

            created_at, link_url, anchor, page_url, da, dr, price, contact, user_first_name, user_last_name, user_email = row_list
            link_url, anchor, page_url = str(link_url).strip(), str(anchor).strip(), str(page_url).strip()

            if not all((page_url, link_url, anchor)):
                validation_errors.append(
                    f'error at row number {num}: page_url/link_url/anchor are mandatory cells')
                continue

            if not link_url.endswith('/'):
                link_url = link_url + '/'

            link_url_domain_name = get_domain_name_from_url(link_url)
            link_url_domain_ser = LinkUrlDomainCreateSerializer(name=link_url_domain_name)
            link_url_domain = get_or_create(db, LinkUrlDomainModel, link_url_domain_ser)

            page_url_domain_name = get_domain_name_from_url(page_url)
            page_url_domain_ser = PageUrlDomainCreateSerializer(name=page_url_domain_name)
            page_url_domain = get_or_create(db, PageUrlDomainModel, page_url_domain_ser)

            user = get(db, UserModel, email=user_email)
            if user is None:
                user = create_user_from_keycloak(db, kc_admin, user_email, user_first_name, user_last_name)

            link_data = {
                'user_id': user.id,
                'created_at': created_at,
                'page_url': page_url,
                'link_url': link_url,
                'anchor': anchor,
                'da': da,
                'dr': dr,
                'price': price,
                'contact': contact,
                'link_url_domain_id': link_url_domain.id,
                'page_url_domain_id': page_url_domain.id,
            }
            try:
                link_create_ser = LinkCreateWithDomainsSerializer(**link_data)
                link_create_ser_list.append(link_create_ser)
            except ValidationError as e:
                error = f'error at row number {num}: {e}'
                validation_errors.append(error)
                print(error)

    if validation_errors:
        if mode is None:
            raise fa.HTTPException(status_code=fa.status.HTTP_422_UNPROCESSABLE_ENTITY,
                                   detail={'validation_errors': validation_errors})
        elif mode == 'from_archive':
            raise WhileUploadingArchiveException(message='validation errors while creating serializers',
                                                 errors=validation_errors,
                                                 user=current_user)
    return link_create_ser_list
