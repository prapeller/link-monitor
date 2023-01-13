import logging

import fastapi as fa
import openpyxl
from pydantic.error_wrappers import ValidationError

from celery_app import celery_app
from core.exceptions import NotExcelException, WhileUploadingArchiveException
from database import SessionLocal
from database.crud import get, create_user_from_keycloak, get_or_create_by_name, get_or_create_many_links_from_archive
from database.models import init_models
from database.models.link_url_domain import LinkUrlDomainModel
from database.models.page_url_domain import PageUrlDomainModel
from database.models.user import UserModel
from database.schemas.link import LinkCreateWithDomainsSerializer
from services.domain_checker.celery_tasks import check_pudomains_with_similarweb
from services.domain_checker.domain_checker import get_domain_name_from_url
from services.keycloak.keycloak import KCAdmin
from services.link_checker.celery_tasks import check_links_from_list

logger = logging.getLogger(name='file_handler')
logger.setLevel(logging.DEBUG)
formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s ")
file_handler = logging.FileHandler(f'services/file_handler/file_handler.log', encoding='utf-8')
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)


def iterfile(filepath):
    with open(filepath, mode="rb") as file:
        yield from file


def get_duplicated_link_row_numbers(ws_links: list[tuple], duplicated_link: tuple) -> str:
    """check if link_tuple(link_url, anchor, page_url) has duplicates in list of such tuples (ws_link_list)"""
    row_numbers = []

    for row_num, link in enumerate(ws_links, start=1):
        if link == duplicated_link:
            row_numbers.append(str(row_num))

    return ', '.join(row_numbers)


def check_file_on_duplicates(filepath) -> list[str]:
    """
    check if file has duplicated links (violating unique constraint (link_url, anchor, page_url))
    and returns list of 'error strings' like
    [
        "file has duplicated links at rows 1, 10",
        "file has duplicated links at rows 2, 12",
    ]
    """
    try:
        wb = openpyxl.load_workbook(filename=filepath)
    except KeyError:
        raise NotExcelException
    ws = wb.active

    ws_links: list[tuple] = []
    ws_duplicated_links: list[tuple] = []
    validation_errors = []

    for row in ws.iter_rows():
        created_at, link_url, anchor, page_url = [cell.value for cell in row[:4]]
        if all((created_at, link_url, anchor, page_url)):
            if not link_url.endswith('/'):
                link_url = link_url + '/'
            ws_links.append((link_url, anchor, page_url))
        else:
            continue

    for link in ws_links:
        if ws_links.count(link) > 1:
            ws_duplicated_links.append(link)

    if ws_duplicated_links:
        ws_duplicated_links = list(set(ws_duplicated_links))

        for duplicated_link in ws_duplicated_links:
            row_numbers = get_duplicated_link_row_numbers(ws_links, duplicated_link)
            error = f'error at row numbers {row_numbers}: link with ' \
                    f'link_url={duplicated_link[0]}, ' \
                    f'anchor={duplicated_link[1]}, ' \
                    f'page_url={duplicated_link[2]} ' \
                    f'is duplicated in file'
            logger.warning(f'check_file_on_duplicates({filepath=:}), appending validation error {error}')
            validation_errors.append(error)

    return validation_errors


def get_link_create_ser_list_from_file(filepath, current_user_id: int, mode=None) \
        -> list[LinkCreateWithDomainsSerializer]:
    session = SessionLocal()
    current_user = get(session, UserModel, id=current_user_id)

    try:
        wb = openpyxl.load_workbook(filename=filepath)
    except KeyError:
        raise NotExcelException
    ws = wb.active

    link_create_ser_list = []
    validation_errors = []

    if mode is None:

        for num, row in enumerate(ws.iter_rows(), start=1):
            if num == 1:
                continue

            row_list = [cell.value for cell in row[:8]]
            if all((cell is None for cell in row_list)):
                continue

            page_url, link_url, anchor, da, dr, price, screenshot_url, contact = row_list
            link_url, anchor, page_url = str(link_url).strip(), str(anchor).strip(), str(page_url).strip()

            if not all((page_url, link_url, anchor, da, dr, price, contact)):
                error = f'error at row number {num}: there are empty cells, only screenshot_url can be empty'
                logger.warning(f'get_link_create_ser_list_from_file({filepath=:}, {current_user_id=:}, {mode=:}), '
                               f'appending validation error {error}')
                validation_errors.append(error)
                continue

            if not link_url.endswith('/'):
                link_url = link_url + '/'

            link_url_domain_name = get_domain_name_from_url(link_url)
            link_url_domain, lud_created = get_or_create_by_name(session, LinkUrlDomainModel, link_url_domain_name)

            page_url_domain_name = get_domain_name_from_url(page_url)
            page_url_domain, pud_created = get_or_create_by_name(session, PageUrlDomainModel, page_url_domain_name)

            check_pudomains_with_similarweb.delay(id_list=[page_url_domain.id])

            link_data = {
                'user_id': current_user_id,
                'page_url': page_url,
                'link_url': link_url,
                'anchor': anchor,
                'da': da,
                'dr': dr,
                'price': price,
                'screenshot_url': screenshot_url,
                'contact': contact,
                'link_url_domain_id': link_url_domain.id,
                'page_url_domain_id': page_url_domain.id,
            }
            try:
                link_create_ser = LinkCreateWithDomainsSerializer(**link_data)
                link_create_ser_list.append(link_create_ser)
            except ValidationError as e:
                error = f'error at row number {num}: {e}'
                logger.warning(f'get_link_create_ser_list_from_file({filepath=:}, {current_user_id=:}, {mode=:}), '
                               f'appending validation error {error}')
                validation_errors.append(error)

    elif mode == 'from_archive':
        kc_admin = KCAdmin()

        for num, row in enumerate(ws.iter_rows(), start=1):
            if num == 1:
                continue

            row_list = [cell.value for cell in row[:11]]
            if all((cell is None for cell in row_list)):
                continue

            created_at, link_url, anchor, page_url, da, dr, price, contact, user_first_name, user_last_name, user_email = row_list
            link_url, anchor, page_url = str(link_url).strip(), str(anchor).strip(), str(page_url).strip()

            if not all((page_url, link_url, anchor)):
                error = f'error at row number {num}: page_url/link_url/anchor are mandatory cells'
                logger.warning(f'get_link_create_ser_list_from_file({filepath=:}, {current_user_id=:}, {mode=:}), '
                               f'appending validation error {error}')
                validation_errors.append(error)
                continue

            if not link_url.endswith('/'):
                link_url = link_url + '/'

            link_url_domain_name = get_domain_name_from_url(link_url)
            link_url_domain, lud_created = get_or_create_by_name(session, LinkUrlDomainModel, link_url_domain_name)

            page_url_domain_name = get_domain_name_from_url(page_url)
            page_url_domain, pud_created = get_or_create_by_name(session, PageUrlDomainModel, page_url_domain_name)

            user = get(session, UserModel, email=user_email)
            if user is None:
                user = create_user_from_keycloak(session, kc_admin, user_email, user_first_name, user_last_name)

            link_data = {
                'user_id': user.id,
                'created_at': created_at,
                'page_url': page_url,
                'link_url': link_url,
                'anchor': anchor,
                'da': da,
                'dr': dr,
                'price': price,
                'contact': contact,
                'link_url_domain_id': link_url_domain.id,
                'page_url_domain_id': page_url_domain.id,
            }
            try:
                link_create_ser = LinkCreateWithDomainsSerializer(**link_data)
                link_create_ser_list.append(link_create_ser)
            except ValidationError as e:
                error = f'error at row number {num}: {e}'
                logger.warning(f'get_link_create_ser_list_from_file({filepath=:}, {current_user_id=:}, {mode=:}), '
                               f'appending validation error {error}')
                validation_errors.append(error)

    if validation_errors:
        if mode is None:
            raise fa.HTTPException(status_code=fa.status.HTTP_422_UNPROCESSABLE_ENTITY,
                                   detail={'validation_errors': validation_errors})
        elif mode == 'from_archive':
            raise WhileUploadingArchiveException(message='validation errors while creating serializers',
                                                 errors=validation_errors,
                                                 user=current_user)
    return link_create_ser_list


@celery_app.task(name='create_links_from_uploaded_file_archive')
def create_links_from_uploaded_file_archive(uploaded_file_path, current_user_id):
    init_models()
    session = SessionLocal()
    current_user = get(session, UserModel, id=current_user_id)
    duplicate_validation_errors = check_file_on_duplicates(filepath=uploaded_file_path)
    if duplicate_validation_errors:
        raise WhileUploadingArchiveException(message='duplicates in archive were found',
                                             errors=duplicate_validation_errors,
                                             user=current_user)
    link_create_ser_list = get_link_create_ser_list_from_file(filepath=uploaded_file_path,
                                                              current_user_id=current_user.id,
                                                              mode='from_archive')
    links = get_or_create_many_links_from_archive(session, link_create_ser_list)
    links_id_list = [str(link.id) for link in links]
    check_links_from_list.delay(id_list=links_id_list)
    session.close()
