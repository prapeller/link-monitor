import copy
import datetime

import openpyxl
import sqlalchemy as sa
from openpyxl.styles import PatternFill
from sqlalchemy.orm import Session

from core.exceptions import NotExcelException
from database.models.link import LinkModel
from database.models.link_url_domain import LinkUrlDomainModel
from database.models.user import UserModel


def get_year_week(date: datetime.datetime):
    """if date month is January and week is 52 (last year week) - then it is 1st week"""
    first_jan_week_number = int(datetime.datetime(date.year, 1, 1).strftime("%V"))
    if date.month == 1 and date.day == 1 and first_jan_week_number == 53:
        return 1
    if first_jan_week_number == 53:
        week_number = int(date.strftime("%V")) + 1
    else:
        week_number = int(date.strftime("%V"))

    if week_number == 52 and date.month == 1:
        return 1
    assert week_number in range(1, 54), f'week number {week_number} should be from 1 to 53'
    return week_number


def get_week_number(date: datetime.datetime):
    """should return only weeks from 1 to 5"""

    first_day_of_month = date.replace(day=1)
    month_starts_in_workweek = False
    if first_day_of_month.weekday() in (0, 1, 2, 3, 4):
        month_starts_in_workweek = True

    first_week_of_month = get_year_week(first_day_of_month)
    date_week = get_year_week(date)

    week_number = date_week - first_week_of_month + 1

    if date.month > 1 and not month_starts_in_workweek and week_number > 1:
        # if date month starts on weekend - next week also will be 1st
        # in case if date month is January all works fine there and no need to subtract 1week
        # (bcz above in get_week() we determined that in such cases first week will be 1st anyway)
        week_number = week_number - 1
    assert week_number in (1, 2, 3, 4, 5), f"week number: {week_number}, for date: {date} should be from 1 to 5"
    return week_number


empty_weeks = {
    '1_1': {'total_count': 0, 'links': []},
    '1_2': {'total_count': 0, 'links': []},
    '1_3': {'total_count': 0, 'links': []},
    '1_4': {'total_count': 0, 'links': []},
    '1_5': {'total_count': 0, 'links': []},
    '2_1': {'total_count': 0, 'links': []},
    '2_2': {'total_count': 0, 'links': []},
    '2_3': {'total_count': 0, 'links': []},
    '2_4': {'total_count': 0, 'links': []},
    '2_5': {'total_count': 0, 'links': []},
    '3_1': {'total_count': 0, 'links': []},
    '3_2': {'total_count': 0, 'links': []},
    '3_3': {'total_count': 0, 'links': []},
    '3_4': {'total_count': 0, 'links': []},
    '3_5': {'total_count': 0, 'links': []},
    '4_1': {'total_count': 0, 'links': []},
    '4_2': {'total_count': 0, 'links': []},
    '4_3': {'total_count': 0, 'links': []},
    '4_4': {'total_count': 0, 'links': []},
    '4_5': {'total_count': 0, 'links': []},
    '5_1': {'total_count': 0, 'links': []},
    '5_2': {'total_count': 0, 'links': []},
    '5_3': {'total_count': 0, 'links': []},
    '5_4': {'total_count': 0, 'links': []},
    '5_5': {'total_count': 0, 'links': []},
    '6_1': {'total_count': 0, 'links': []},
    '6_2': {'total_count': 0, 'links': []},
    '6_3': {'total_count': 0, 'links': []},
    '6_4': {'total_count': 0, 'links': []},
    '6_5': {'total_count': 0, 'links': []},
    '7_1': {'total_count': 0, 'links': []},
    '7_2': {'total_count': 0, 'links': []},
    '7_3': {'total_count': 0, 'links': []},
    '7_4': {'total_count': 0, 'links': []},
    '7_5': {'total_count': 0, 'links': []},
    '8_1': {'total_count': 0, 'links': []},
    '8_2': {'total_count': 0, 'links': []},
    '8_3': {'total_count': 0, 'links': []},
    '8_4': {'total_count': 0, 'links': []},
    '8_5': {'total_count': 0, 'links': []},
    '9_1': {'total_count': 0, 'links': []},
    '9_2': {'total_count': 0, 'links': []},
    '9_3': {'total_count': 0, 'links': []},
    '9_4': {'total_count': 0, 'links': []},
    '9_5': {'total_count': 0, 'links': []},
    '10_1': {'total_count': 0, 'links': []},
    '10_2': {'total_count': 0, 'links': []},
    '10_3': {'total_count': 0, 'links': []},
    '10_4': {'total_count': 0, 'links': []},
    '10_5': {'total_count': 0, 'links': []},
    '11_1': {'total_count': 0, 'links': []},
    '11_2': {'total_count': 0, 'links': []},
    '11_3': {'total_count': 0, 'links': []},
    '11_4': {'total_count': 0, 'links': []},
    '11_5': {'total_count': 0, 'links': []},
    '12_1': {'total_count': 0, 'links': []},
    '12_2': {'total_count': 0, 'links': []},
    '12_3': {'total_count': 0, 'links': []},
    '12_4': {'total_count': 0, 'links': []},
    '12_5': {'total_count': 0, 'links': []},
}


def generate_report(
        db: Session,
        template_filename: str,
        to_save_filepath: str,
        link_url_domains: list[LinkUrlDomainModel],
        users: list[UserModel]
):
    current_year = datetime.datetime.now().year
    users_id_list = [user.id for user in users]

    try:
        wb = openpyxl.load_workbook(filename=template_filename)
    except KeyError:
        raise NotExcelException
    ws = wb.active

    current_row = 3

    for link_url_domain in link_url_domains:
        ws.append([link_url_domain.name])
        current_row += 1

        domain_acceptors: list[LinkModel] = db.query(LinkModel).filter(
            LinkModel.link_url_domain_id == link_url_domain.id).filter(
            LinkModel.user_id.in_(users_id_list)).distinct(
            LinkModel.link_url).all()

        for acceptor in domain_acceptors:
            ws.append(['', acceptor.link_url])
            current_row += 1

            acceptor_links = db.query(LinkModel).filter(LinkModel.link_url == acceptor.link_url).all()
            acceptor_links = [link for link in acceptor_links if link.created_at.year == current_year]

            acceptor_users: set[UserModel] = set(users).intersection(set([link.user for link in acceptor_links]))

            if acceptor_users:
                acceptor_users_links_sum_months = [
                    [0, '', '', '', ''],
                    [0, '', '', '', ''],
                    [0, '', '', '', ''],
                    [0, '', '', '', ''],
                    [0, '', '', '', ''],
                    [0, '', '', '', ''],
                    [0, '', '', '', ''],
                    [0, '', '', '', ''],
                    [0, '', '', '', ''],
                    [0, '', '', '', ''],
                    [0, '', '', '', ''],
                    [0, '', '', '', ''],
                ]

                for user in acceptor_users:
                    ws.append(['', '', f"{user.first_name} {user.last_name}"])
                    current_row += 1

                    user_acceptor_links: list[LinkModel] = [link for link in acceptor_links if link.user_id == user.id]

                    for link in user_acceptor_links:
                        anchor = link.anchor
                        page_url = link.page_url
                        linkcheck_last_status = link.link_check_last_status
                        linkcheck_last_result = link.link_check_last_result_message
                        created_at_month = link.created_at.month
                        acceptor_users_links_sum_months[created_at_month - 1][0] += 1

                        empty_months_row = ['' for _ in range(1, (created_at_month - 1) * 5)]
                        created_at_month_row = ['' for _ in range(get_week_number(link.created_at))] + [1]

                        ws.append(['', '', '', anchor, page_url, linkcheck_last_status, linkcheck_last_result,
                                   *empty_months_row, *created_at_month_row])
                        current_row += 1

                        fill = PatternFill(fill_type='solid', start_color='e3e3e3', end_color='e3e3e3')
                        for col in range(4, 68):
                            cell = ws.cell(row=current_row, column=col)
                            cell.fill = fill

                sum_acceptor_users_links_row = ['' for _ in range(7)]
                for month in acceptor_users_links_sum_months:
                    sum_acceptor_users_links_row.extend(month)
                ws.append(sum_acceptor_users_links_row)
                current_row += 1

            fill = PatternFill(fill_type='solid', start_color='c3c3c3', end_color='c3c3c3')
            for col in range(3, 68):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = fill
    wb.save(to_save_filepath)


def generate_report_v2(
        db: Session,
        template_filename: str,
        to_save_filepath: str,
        link_url_domains: list[LinkUrlDomainModel],
        users: list[UserModel],
        year: int,
):
    """same as above but based on particular year"""
    users_id_list = [user.id for user in users]

    try:
        wb = openpyxl.load_workbook(filename=template_filename)
    except KeyError:
        raise NotExcelException
    ws = wb.active

    current_row = 3

    for link_url_domain in link_url_domains:
        ws.append([link_url_domain.name])
        current_row += 1

        domain_acceptors: list[LinkModel] = db.query(LinkModel).filter(
            LinkModel.link_url_domain_id == link_url_domain.id).filter(
            LinkModel.user_id.in_(users_id_list)).distinct(
            LinkModel.link_url).all()

        for acceptor in domain_acceptors:
            ws.append(['', acceptor.link_url])
            current_row += 1

            acceptor_links = db.query(LinkModel).filter(LinkModel.link_url == acceptor.link_url).all()
            acceptor_links = [link for link in acceptor_links if link.created_at.year == year]

            acceptor_users: set[UserModel] = set(users).intersection(set([link.user for link in acceptor_links]))

            if acceptor_users:
                acceptor_users_links_sum_months = [
                    [0, '', '', '', ''],
                    [0, '', '', '', ''],
                    [0, '', '', '', ''],
                    [0, '', '', '', ''],
                    [0, '', '', '', ''],
                    [0, '', '', '', ''],
                    [0, '', '', '', ''],
                    [0, '', '', '', ''],
                    [0, '', '', '', ''],
                    [0, '', '', '', ''],
                    [0, '', '', '', ''],
                    [0, '', '', '', ''],
                ]

                for user in acceptor_users:
                    ws.append(['', '', f"{user.first_name} {user.last_name}"])
                    current_row += 1

                    user_acceptor_links: list[LinkModel] = [link for link in acceptor_links if link.user_id == user.id]

                    for link in user_acceptor_links:
                        anchor = link.anchor
                        page_url = link.page_url
                        linkcheck_last_status = link.link_check_last_status
                        linkcheck_last_result = link.link_check_last_result_message
                        created_at_month = link.created_at.month
                        acceptor_users_links_sum_months[created_at_month - 1][0] += 1

                        empty_months_row = ['' for _ in range(1, (created_at_month - 1) * 5)]
                        created_at_month_row = ['' for _ in range(get_week_number(link.created_at))] + [1]

                        ws.append(['', '', '', anchor, page_url, linkcheck_last_status, linkcheck_last_result,
                                   *empty_months_row, *created_at_month_row])
                        current_row += 1

                        fill = PatternFill(fill_type='solid', start_color='e3e3e3', end_color='e3e3e3')
                        for col in range(4, 68):
                            cell = ws.cell(row=current_row, column=col)
                            cell.fill = fill

                sum_acceptor_users_links_row = ['' for _ in range(7)]
                for month in acceptor_users_links_sum_months:
                    sum_acceptor_users_links_row.extend(month)
                ws.append(sum_acceptor_users_links_row)
                current_row += 1

            fill = PatternFill(fill_type='solid', start_color='c3c3c3', end_color='c3c3c3')
            for col in range(3, 68):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = fill
    wb.save(to_save_filepath)


def generate_report_ui(
        db: Session,
        link_url_domains: list[LinkUrlDomainModel],
        users=list[UserModel]
):
    result = {}
    current_year = datetime.datetime.now().year
    users_id_list = [user.id for user in users]

    for link_url_domain in link_url_domains:
        result[f'{link_url_domain.name}'] = {}

        domain_acceptors: list[LinkModel] = db.query(LinkModel) \
            .filter(LinkModel.link_url_domain_id == link_url_domain.id) \
            .filter(sa.extract('year', LinkModel.created_at) == current_year) \
            .distinct(LinkModel.link_url).all()

        for acceptor in domain_acceptors:
            result[f'{link_url_domain.name}'][f'{acceptor.link_url}'] = copy.deepcopy(empty_weeks)

            acceptor_users = db.query(UserModel).join(LinkModel) \
                .filter(UserModel.id.in_(users_id_list)) \
                .filter(LinkModel.link_url == acceptor.link_url) \
                .all()

            for user in acceptor_users:
                user_acceptor_links = db.query(LinkModel) \
                    .filter(LinkModel.link_url == acceptor.link_url) \
                    .filter(sa.extract('year', LinkModel.created_at) == current_year) \
                    .filter(LinkModel.user_id == user.id) \
                    .all()
                for link in user_acceptor_links:
                    anchor = link.anchor
                    page_url = link.page_url
                    linkcheck_last_status = link.link_check_last_status
                    linkcheck_last_result = link.link_check_last_result_message
                    created_at_month = link.created_at.month
                    created_at_week = get_week_number(link.created_at)
                    created_at_month_week = f'{created_at_month}_{created_at_week}'
                    link_dict = {
                        'id': link.id,
                        'user': str(user),
                        'anchor': anchor,
                        'page_url': page_url,
                        'linkcheck_last_status': linkcheck_last_status,
                        'linkcheck_last_result': linkcheck_last_result,
                        'created_at_month_week': created_at_month_week,
                    }
                    result[f'{link_url_domain.name}'][f'{acceptor.link_url}'][created_at_month_week][
                        'total_count'] += 1
                    result[f'{link_url_domain.name}'][f'{acceptor.link_url}'][created_at_month_week][
                        'links'].append(link_dict)

    return result


def generate_report_ui_v2(
        db: Session,
        link_url_domains: list[LinkUrlDomainModel],
        users=list[UserModel],
        year=int,
):
    """same as above, but based on particular year"""
    result = {}
    users_id_list = [user.id for user in users]

    for link_url_domain in link_url_domains:
        result[f'{link_url_domain.name}'] = {}

        start_time = datetime.datetime.now()

        domain_acceptors: list[LinkModel] = db.query(LinkModel) \
            .filter(LinkModel.link_url_domain_id == link_url_domain.id) \
            .filter(sa.extract('year', LinkModel.created_at) == year) \
            .distinct(LinkModel.link_url).all()

        for acceptor in domain_acceptors:
            result[f'{link_url_domain.name}'][f'{acceptor.link_url}'] = copy.deepcopy(empty_weeks)

            acceptor_users = db.query(UserModel).join(LinkModel) \
                .filter(LinkModel.link_url == acceptor.link_url) \
                .filter(UserModel.id.in_(users_id_list)) \
                .all()

            for user in acceptor_users:

                user_acceptor_links = db.query(LinkModel) \
                    .filter(LinkModel.link_url == acceptor.link_url) \
                    .filter(sa.extract('year', LinkModel.created_at) == year) \
                    .filter(LinkModel.user_id == user.id) \
                    .all()

                for link in user_acceptor_links:
                    anchor = link.anchor
                    page_url = link.page_url
                    linkcheck_last_status = link.link_check_last_status
                    linkcheck_last_result = link.link_check_last_result_message
                    created_at_month = link.created_at.month
                    created_at_week = get_week_number(link.created_at)
                    created_at_month_week = f'{created_at_month}_{created_at_week}'
                    link_dict = {
                        'id': link.id,
                        'user': str(user),
                        'anchor': anchor,
                        'page_url': page_url,
                        'linkcheck_last_status': linkcheck_last_status,
                        'linkcheck_last_result': linkcheck_last_result,
                        'created_at_month_week': created_at_month_week,
                    }
                    result[f'{link_url_domain.name}'][f'{acceptor.link_url}'][created_at_month_week][
                        'total_count'] += 1
                    result[f'{link_url_domain.name}'][f'{acceptor.link_url}'][created_at_month_week][
                        'links'].append(link_dict)
        finish_time = datetime.datetime.now()
        print(finish_time - start_time)

    return result
